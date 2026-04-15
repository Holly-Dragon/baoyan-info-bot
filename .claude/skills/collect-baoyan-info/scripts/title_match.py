import openpyxl
import requests
import re
import urllib3
import os

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def update_titles(excel_file="2026院校信息_更新.xlsx"):
    if not os.path.exists(excel_file):
        print(f"文件 {excel_file} 不存在，已跳过标题匹配。")
        return

    print("开始匹配通知标题并更新“招生项目”字段...")
    
    wb = openpyxl.load_workbook(excel_file)
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 查找“通知链接”和“招生项目”所在的列索引
        link_col_idx = None
        project_col_idx = None
        
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value == "通知链接":
                link_col_idx = col
            elif header_value == "招生项目":
                project_col_idx = col
                
        if not link_col_idx or not project_col_idx:
            continue
            
        for row in range(2, ws.max_row + 1):
            link_cell = ws.cell(row=row, column=link_col_idx)
            link_function = str(link_cell.value).strip() if link_cell.value else ""
            link = link_function.strip('=HYPERLINK("').split('",')[0] if link_function.startswith('=HYPERLINK("') else link_function
            if not link.startswith("http"):
                continue
                
            try:
                res = requests.get(link, headers=headers, timeout=5, verify=False)
                res.encoding = 'utf-8'
                
                title = None
                if link.startswith("https://mp.weixin.qq.com"):
                    match = re.search(r'<meta property="og:title"\s+content="(.*?)"\s*/>', res.text, re.IGNORECASE)
                    if match:
                        title = match.group(1).replace('\n', '').strip()
                else:
                    match = re.search(r'<title>(.*?)</title>', res.text, re.IGNORECASE)
                    if match:
                        title = match.group(1).replace('\n', '').strip()
                
                if title:
                    print(f"[{sheet_name} 行 {row}] 更新标题为: {title}")
                    ws.cell(row=row, column=project_col_idx).value = title

            except Exception as e:
                print(f"[{sheet_name} 行 {row}] 抓取失败: {e}")

    # 直接保存 workbook，完美保留所有原有的Excel公式与样式
    wb.save(excel_file)
    print("所有标签页标题匹配更新完成！公式和格式已保留。")

if __name__ == "__main__":
    update_titles()
