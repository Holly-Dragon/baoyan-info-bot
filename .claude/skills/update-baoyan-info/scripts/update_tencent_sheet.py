#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将本地 Excel 数据追加写入腾讯共享文档

功能：
1. 读取本地 Excel 文件 (2026院校信息_更新.xlsx)
2. 获取腾讯在线表格的现有数据
3. 对比去重，找出新增数据
4. 将新增数据追加写入在线表格
5. "申请倒计时"字段使用公式实时计算

使用方式：
    python update_tencent_sheet.py [--excel EXCEL_PATH] [--url SHEET_URL]
    python update_tencent_sheet.py --check-env  # 检查环境配置

依赖：
    - openpyxl: 读取本地 Excel
    - mcporter: 调用腾讯文档 API (需要配置好 PATH 或通过参数指定)

环境变量：
    - MCPORTER_PATH: mcporter 可执行文件路径（可选，默认自动检测）
    - NODE_PATH: Node.js bin 目录路径（可选，默认自动检测）
"""

import os
import re
import sys
import json
import subprocess
import argparse
import shutil
from datetime import datetime
from typing import List, Dict, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("错误：缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)


# ============== 配置 ==============

# 默认 Excel 文件路径
DEFAULT_EXCEL_PATH = "2026院校信息_更新.xlsx"

# 默认腾讯文档 URL：https://docs.qq.com/sheet/DWkpjRkZodUF3UG92
# 测试文档 URL：https://docs.qq.com/sheet/DWm5ucU5WY1hsY3Fv
DEFAULT_SHEET_URL = "https://docs.qq.com/sheet/DWkpjRkZodUF3UG92"

# Sheet 名称与 ID 的映射（需要先通过 get_sheet_info 获取）
SHEET_NAME_TO_ID = {
    "理工农医": "000001",
    "经管法": "000002",
    "人文社科与艺术": "000003",
    "单校通知": "000004",
}


# ============== 环境检测 ==============

def detect_node_path() -> Optional[str]:
    """自动检测 Node.js bin 目录路径"""
    # 1. 优先检查环境变量
    env_node_path = os.environ.get("NODE_PATH")
    if env_node_path and os.path.exists(env_node_path):
        node_exe = os.path.join(env_node_path, "node")
        if os.path.exists(node_exe):
            return env_node_path
    
    # 2. 检查常见路径
    common_paths = [
        # NVM 路径（macOS/Linux）
        os.path.expanduser("~/.nvm/versions/node"),
        # Volta 路径
        os.path.expanduser("~/.volta/tools/image/node"),
        # Homebrew 路径（macOS）
        "/opt/homebrew/bin",
        "/usr/local/bin",
        # Windows 路径
        os.path.join(os.environ.get("APPDATA", ""), "npm"),
    ]
    
    for base_path in common_paths:
        if not os.path.exists(base_path):
            continue
        
        # 如果是 nvm 目录，需要找子目录
        if ".nvm" in base_path or ".volta" in base_path:
            try:
                subdirs = sorted(os.listdir(base_path), reverse=True)
                for subdir in subdirs:
                    bin_path = os.path.join(base_path, subdir, "bin")
                    node_exe = os.path.join(bin_path, "node")
                    if os.path.exists(node_exe):
                        return bin_path
            except OSError:
                continue
        else:
            node_exe = os.path.join(base_path, "node")
            if os.path.exists(node_exe):
                return base_path
    
    # 3. 使用 which/where 查找
    node_exe = shutil.which("node")
    if node_exe:
        return os.path.dirname(node_exe)
    
    return None


def detect_mcporter_path() -> Optional[str]:
    """自动检测 mcporter 可执行文件路径"""
    # 1. 优先检查环境变量
    env_mcporter = os.environ.get("MCPORTER_PATH")
    if env_mcporter and os.path.exists(env_mcporter):
        return env_mcporter
    
    # 2. 使用 which/where 查找
    mcporter = shutil.which("mcporter")
    if mcporter:
        return mcporter
    
    # 3. 在 Node.js bin 目录中查找
    node_path = detect_node_path()
    if node_path:
        mcporter_path = os.path.join(node_path, "mcporter")
        if os.path.exists(mcporter_path):
            return mcporter_path
    
    return None


def check_environment() -> Tuple[bool, str, str]:
    """
    检查环境配置
    返回: (是否成功, node_path, mcporter_path)
    """
    print("\n" + "=" * 50)
    print("环境检查")
    print("=" * 50)
    
    # 检查 Node.js
    node_path = detect_node_path()
    if node_path:
        node_exe = os.path.join(node_path, "node")
        try:
            result = subprocess.run(
                [node_exe, "--version"],
                capture_output=True,
                text=True,
                timeout=5
            )
            print(f"✅ Node.js: {result.stdout.strip()} (路径: {node_path})")
        except Exception as e:
            print(f"⚠️  Node.js 路径存在但执行失败: {e}")
            node_path = None
    else:
        print("❌ Node.js: 未找到")
        print("   请安装 Node.js 或设置 NODE_PATH 环境变量")
    
    # 检查 mcporter
    mcporter_path = detect_mcporter_path()
    if mcporter_path:
        print(f"✅ mcporter: {mcporter_path}")
    else:
        print("❌ mcporter: 未找到")
        print("   请安装 mcporter: npm install -g mcporter")
        print("   或设置 MCPORTER_PATH 环境变量")
    
    # 检查 openpyxl
    print(f"✅ openpyxl: 已安装")
    
    success = node_path is not None and mcporter_path is not None
    
    if not success:
        print("\n" + "-" * 50)
        print("⚠️  环境配置不完整，请按以下步骤修复：")
        print("-" * 50)
        if not node_path:
            print("1. 安装 Node.js:")
            print("   macOS: brew install node")
            print("   或使用 nvm: curl -o- https://raw.githubusercontent.com/nvm-sh/nvm/v0.39.0/install.sh | bash")
            print("   Windows: 从 https://nodejs.org 下载安装")
        if not mcporter_path:
            print("2. 安装 mcporter:")
            print("   npm install -g mcporter")
        print("\n3. 设置环境变量（可选，用于自定义路径）:")
        print("   export NODE_PATH=/path/to/node/bin")
        print("   export MCPORTER_PATH=/path/to/mcporter")
        print("-" * 50)
    else:
        print("\n✅ 环境检查通过！")
    
    return success, node_path or "", mcporter_path or ""


# ============== API 封装 ==============

class TencentSheetAPI:
    """腾讯文档 Sheet API 封装"""
    
    def __init__(self, node_path: str = None, mcporter_path: str = None):
        self.node_path = node_path or detect_node_path() or ""
        self.mcporter_path = mcporter_path or detect_mcporter_path() or ""
        
        if not self.mcporter_path:
            raise RuntimeError(
                "未找到 mcporter，请先安装:\n"
                "  npm install -g mcporter\n"
                "或设置环境变量:\n"
                "  export MCPORTER_PATH=/path/to/mcporter"
            )
    
    def _run_mcporter(self, tool: str, args: dict) -> dict:
        """调用 mcporter 执行腾讯文档 API"""
        env = os.environ.copy()
        if self.node_path:
            env["PATH"] = f"{self.node_path}:{env.get('PATH', '')}"
        
        cmd = [
            self.mcporter_path,
            "call",
            "tencent-sheetengine",
            tool,
            "--args",
            json.dumps(args, ensure_ascii=False)
        ]
        
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                env=env,
                timeout=60
            )
            if result.returncode != 0:
                return {"error": result.stderr}
            return json.loads(result.stdout)
        except subprocess.TimeoutExpired:
            return {"error": "timeout"}
        except json.JSONDecodeError as e:
            return {"error": f"json decode error: {e}"}
        except Exception as e:
            return {"error": str(e)}
    
    def get_sheet_info(self, file_url: str) -> List[dict]:
        """获取在线表格的子表信息"""
        result = self._run_mcporter("get_sheet_info", {"file_url": file_url})
        if "error" in result:
            print(f"获取子表信息失败: {result['error']}")
            return []
        return result.get("sheets", [])
    
    def get_cell_data(self, file_url: str, sheet_id: str, 
                      start_row: int, start_col: int, 
                      end_row: int, end_col: int, 
                      return_csv: bool = False) -> dict:
        """获取单元格数据"""
        args = {
            "file_url": file_url,
            "sheet_id": sheet_id,
            "start_row": start_row,
            "start_col": start_col,
            "end_row": end_row,
            "end_col": end_col,
            "return_csv": return_csv
        }
        return self._run_mcporter("get_cell_data", args)
    
    def set_range_value(self, file_url: str, sheet_id: str, values: List[dict]) -> dict:
        """批量设置单元格值"""
        args = {
            "file_url": file_url,
            "sheet_id": sheet_id,
            "values": values
        }
        return self._run_mcporter("set_range_value", args)

    def set_link(self, file_url: str, sheet_id: str, row: int, col: int,
                 url: str, display_text: str) -> dict:
        """为指定单元格设置超链接"""
        args = {
            "file_url": file_url,
            "sheet_id": sheet_id,
            "row": row,
            "col": col,
            "url": url,
            "display_text": display_text
        }
        return self._run_mcporter("set_link", args)


# ============== 数据处理 ==============

def read_local_excel(excel_path: str) -> Dict[str, List[dict]]:
    """读取本地 Excel 文件，返回按 sheet 分类的数据列表"""
    if not os.path.exists(excel_path):
        print(f"Excel 文件不存在: {excel_path}")
        return {}
    
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    result = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row <= 1:
            result[sheet_name] = []
            continue
        
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row_data = {}
            for c, header in enumerate(headers):
                if header:
                    val = ws.cell(r, c + 1).value
                    if header == "通知链接" and isinstance(val, str):
                        import re
                        match = re.search(r'=HYPERLINK\("([^"]+)"', val, re.IGNORECASE)
                        if match:
                            val = match.group(1)
                    row_data[header] = val
            rows.append(row_data)
        
        result[sheet_name] = rows
        print(f"  读取 [{sheet_name}]: {len(rows)} 条数据")
    
    return result


def get_online_data(api: TencentSheetAPI, file_url: str, sheet_id: str) -> Tuple[List[str], List[dict]]:
    """获取在线表格的现有数据"""
    # 共享表格已删除“是否截止”列，仅拉取到“申请倒计时”(第9列，0-based col=8)
    result = api.get_cell_data(file_url, sheet_id, 0, 0, 199, 8, return_csv=True)
    
    if "error" in result:
        print(f"获取在线数据失败: {result['error']}")
        return [], []
    
    csv_data = result.get("csv_data", "")
    if not csv_data:
        return [], []
    
    lines = csv_data.strip().split("\n")
    if not lines:
        return [], []
    
    def parse_csv_line(line: str) -> List[str]:
        fields = []
        current = ""
        in_quotes = False
        for char in line:
            if char == '"':
                in_quotes = not in_quotes
            elif char == ',' and not in_quotes:
                fields.append(current)
                current = ""
            else:
                current += char
        fields.append(current)
        return fields
    
    # 约定：第1行为说明/占位，第2行为字段名，第3行起为实际数据
    if len(lines) < 2:
        return [], []

    headers = parse_csv_line(lines[1])
    data_rows = []
    
    for line in lines[2:]:
        if not line.strip() or line.count(',') < 5:
            continue
        fields = parse_csv_line(line)
        school = fields[2] if len(fields) > 2 else ""
        project = fields[5] if len(fields) > 5 else ""
        if not school.strip() and not project.strip():
            continue
        row = {headers[i]: fields[i] if i < len(fields) else "" for i in range(len(headers))}
        data_rows.append(row)
    
    return headers, data_rows


def find_new_rows(local_rows: List[dict], online_rows: List[dict]) -> List[dict]:
    """对比本地数据和在线数据，找出需要新增的行"""
    online_links = set()
    for row in online_rows:
        link = row.get("通知链接", "")
        if link:
            online_links.add(link)
    
    new_rows = []
    for row in local_rows:
        link = row.get("通知链接", "")
        if link and link not in online_links:
            new_rows.append(row)
    
    return new_rows


def build_cell_values(row: dict, row_index: int) -> Tuple[List[dict], Optional[dict]]:
    """
    构建单元格值列表，包含公式。
    返回 (values, link_info)，其中 link_info 为 {"row": ..., "col": ..., "url": ..., "display": ...} 或 None。
    """
    values = []
    """构建单元格值列表，包含公式"""
    values = []

    def normalize_update_time(raw_value) -> str:
        """统一输出为 2026年X月X日，避免时分秒或其他格式混入共享表格。"""
        if raw_value is None:
            today = datetime.now()
            return f"2026年{today.month}月{today.day}日"

        value = str(raw_value).strip()
        if not value:
            today = datetime.now()
            return f"2026年{today.month}月{today.day}日"

        # 兼容 "2026年4月11日" 这类格式
        zh_match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", value)
        if zh_match:
            month = int(zh_match.group(2))
            day = int(zh_match.group(3))
            return f"2026年{month}月{day}日"

        # 兼容 "2026-04-11" / "2026/04/11" / 带时分秒等格式
        try:
            dt = datetime.fromisoformat(value.replace("/", "-").replace("Z", ""))
            return f"2026年{dt.month}月{dt.day}日"
        except ValueError:
            pass

        ymd_match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", value)
        if ymd_match:
            month = int(ymd_match.group(2))
            day = int(ymd_match.group(3))
            return f"2026年{month}月{day}日"

        today = datetime.now()
        return f"2026年{today.month}月{today.day}日"
    
    field_mapping = [
        ("更新时间", "STRING"),
        ("院校报名信息发布时间", "STRING"),
        ("学校", "STRING"),
        ("学院", "STRING"),
        ("专业", "STRING"),
        ("招生项目", "STRING"),
        ("通知链接", "STRING"),
        ("申请截止时间", "STRING"),
    ]
    
    for col_idx, (field_name, value_type) in enumerate(field_mapping):
        value = row.get(field_name, "")
        if field_name == "更新时间":
            value = normalize_update_time(value)
        if value is None:
            value = ""
        values.append({
            "row": row_index,
            "col": col_idx,
            "value_type": value_type,
            "string_value": str(value)
        })
    
    # 申请倒计时公式
    countdown_formula = f'=IFERROR(IF(DATEVALUE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(H{row_index+1},"年","-"),"月","-"),"日",""))<TODAY(),"已截止",DATEVALUE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(H{row_index+1},"年","-"),"月","-"),"日",""))-TODAY()&"天"),"未知")'
    values.append({
        "row": row_index,
        "col": 8,
        "value_type": "FORMULA",
        "formula": countdown_formula
    })

    # 收集通知链接超链接信息（第6列，0-based）
    link_url = row.get("通知链接", "")
    if link_url and str(link_url).strip():
        link_info = {
            "row": row_index,
            "col": 6,
            "url": str(link_url).strip(),
            "display": str(link_url).strip()
        }
    else:
        link_info = None

    return values, link_info


def append_data_to_sheet(api: TencentSheetAPI, file_url: str, sheet_id: str,
                         new_rows: List[dict], existing_row_count: int) -> int:
    """将新数据追加写入在线表格"""
    if not new_rows:
        return 0

    start_row = existing_row_count
    all_values = []
    all_link_info = []

    for i, row in enumerate(new_rows):
        row_index = start_row + i
        values, link_info = build_cell_values(row, row_index)
        all_values.extend(values)
        if link_info:
            all_link_info.append(link_info)

    result = api.set_range_value(file_url, sheet_id, all_values)

    if "error" in result:
        print(f"  写入失败: {result['error']}")
        return 0

    # 单独写入超链接（通知链接列）
    for link in all_link_info:
        link_result = api.set_link(
            file_url=file_url,
            sheet_id=sheet_id,
            row=link["row"],
            col=link["col"],
            url=link["url"],
            display_text=link["display"]
        )
        if "error" in link_result:
            print(f"  ⚠️  超链接写入失败 (行 {link['row']}, 列 {link['col']}): {link_result['error']}")

    return len(new_rows)


# ============== 主流程 ==============

def update_tencent_sheet(excel_path: str, sheet_url: str, 
                         node_path: str = None, mcporter_path: str = None) -> dict:
    """主函数：更新腾讯共享表格"""
    print(f"\n{'='*50}")
    print(f"开始更新腾讯共享表格")
    print(f"{'='*50}")
    print(f"本地 Excel: {excel_path}")
    print(f"在线表格: {sheet_url}")
    
    # 初始化 API
    try:
        api = TencentSheetAPI(node_path, mcporter_path)
    except RuntimeError as e:
        print(f"\n❌ {e}")
        return {"success": False, "message": str(e)}
    
    # 1. 读取本地 Excel
    print(f"\n[1/4] 读取本地 Excel...")
    local_data = read_local_excel(excel_path)
    if not local_data:
        print("本地 Excel 无数据，退出。")
        return {"success": False, "message": "本地 Excel 无数据"}
    
    total_local_rows = sum(len(rows) for rows in local_data.values())
    print(f"本地 Excel 共 {total_local_rows} 条数据")
    
    # 2. 获取在线表格信息
    print(f"\n[2/4] 获取在线表格信息...")
    sheets_info = api.get_sheet_info(sheet_url)
    if not sheets_info:
        print("获取在线表格信息失败，退出。")
        return {"success": False, "message": "获取在线表格信息失败"}
    
    for sheet in sheets_info:
        SHEET_NAME_TO_ID[sheet["sheet_name"]] = sheet["sheet_id"]
    # 别名映射："单校" 在在线表格中名为"单校通知"（ID: 56j4fo）
    if SHEET_NAME_TO_ID.get("单校") in ("000004", None) and "单校通知" in SHEET_NAME_TO_ID:
        SHEET_NAME_TO_ID["单校"] = SHEET_NAME_TO_ID["单校通知"]
    print(f"子表映射: {SHEET_NAME_TO_ID}")
    
    # 3. 对比并找出新增数据
    print(f"\n[3/4] 对比数据，找出新增记录...")
    stats = {}
    
    for sheet_name, local_rows in local_data.items():
        sheet_id = SHEET_NAME_TO_ID.get(sheet_name)
        if not sheet_id:
            print(f"  跳过 [{sheet_name}]: 在线表格中无对应子表")
            continue
        
        print(f"\n  处理 [{sheet_name}]...")
        
        _, online_rows = get_online_data(api, sheet_url, sheet_id)
        print(f"    在线数据: {len(online_rows)} 条")
        print(f"    本地数据: {len(local_rows)} 条")
        
        new_rows = find_new_rows(local_rows, online_rows)
        print(f"    新增数据: {len(new_rows)} 条")
        
        stats[sheet_name] = {
            "online": len(online_rows),
            "local": len(local_rows),
            "new": len(new_rows)
        }
    
    # 4. 写入新增数据
    print(f"\n[4/4] 写入新增数据...")
    total_written = 0
    
    for sheet_name, local_rows in local_data.items():
        sheet_id = SHEET_NAME_TO_ID.get(sheet_name)
        if not sheet_id:
            continue
        
        stat = stats.get(sheet_name, {})
        new_count = stat.get("new", 0)
        if new_count == 0:
            print(f"  [{sheet_name}]: 无新增数据，跳过")
            continue
        
        _, online_rows = get_online_data(api, sheet_url, sheet_id)
        # 共享表格约定：第1行跳过，第2行为表头，第3行开始写数据（0-based 起始行为 2）
        existing_row_count = len(online_rows) + 2
        
        new_rows = find_new_rows(local_rows, online_rows)
        
        written = append_data_to_sheet(api, sheet_url, sheet_id, new_rows, existing_row_count)
        total_written += written
        print(f"  [{sheet_name}]: 成功写入 {written} 条")
    
    # 5. 汇总
    print(f"\n{'='*50}")
    print(f"更新完成！")
    print(f"{'='*50}")
    print(f"总计写入: {total_written} 条新记录")
    for sheet_name, stat in stats.items():
        print(f"  [{sheet_name}]: 在线 {stat['online']} 条 + 新增 {stat['new']} 条")
    
    return {
        "success": True,
        "total_written": total_written,
        "stats": stats
    }


def main():
    parser = argparse.ArgumentParser(
        description="更新腾讯共享表格",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 检查环境配置
  python update_tencent_sheet.py --check-env

  # 使用默认配置更新表格
  python update_tencent_sheet.py

  # 指定 Excel 文件和表格 URL
  python update_tencent_sheet.py --excel data.xlsx --url https://docs.qq.com/sheet/xxx

  # 指定 mcporter 路径（跨设备使用时）
  python update_tencent_sheet.py --mcporter /path/to/mcporter

环境变量:
  NODE_PATH      Node.js bin 目录路径
  MCPORTER_PATH  mcporter 可执行文件路径
        """
    )
    parser.add_argument("--excel", default=DEFAULT_EXCEL_PATH, help="本地 Excel 文件路径")
    parser.add_argument("--url", default=DEFAULT_SHEET_URL, help="腾讯文档 URL")
    parser.add_argument("--check-env", action="store_true", help="检查环境配置")
    parser.add_argument("--node-path", help="Node.js bin 目录路径")
    parser.add_argument("--mcporter", help="mcporter 可执行文件路径")
    args = parser.parse_args()
    
    # 环境检查模式
    if args.check_env:
        success, node_path, mcporter_path = check_environment()
        sys.exit(0 if success else 1)
    
    # 支持相对路径
    excel_path = args.excel
    if not os.path.isabs(excel_path):
        excel_path = os.path.join(os.getcwd(), excel_path)
    
    result = update_tencent_sheet(
        excel_path, 
        args.url,
        node_path=args.node_path,
        mcporter_path=args.mcporter
    )
    
    if not result["success"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
